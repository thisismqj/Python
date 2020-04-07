import urllib.request
import re
from openpyxl import load_workbook

global word
word = ""
global piece
piece = ""
global htm
htm = ""
global speech
global means
rule = r"<ul class=\"dict-basic-ul\">.*?</ul>"
rule1 = r"(?<=<span>).*?(?=</span>)"
rule2 = r"(?<=<strong>).*?(?=</strong>)"
rule3 = r"[^；]*"
def boot():
    global htm
    global piece
    htm = urllib.request.urlopen("http://dict.cn/"+word).read().decode("utf-8")
    piece = re.search(rule,htm,re.DOTALL)[0]
def cspeech():
    global piece
    speech = re.findall(rule1, piece, re.DOTALL)
    return speech
def cmeans():
    global means
    means = []
    for mean in re.findall(rule2, piece):
        temp = []
        for unit in re.findall(rule3, mean):
            if unit:
                temp.append(unit)
        means.append(temp)
    return means

wdb = load_workbook("wb2.xlsx")
wd = wdb["Base"]

this = 2297
for time in range(39):
    word = wd.cell(this, 1).value
    boot()
    sp = cspeech()
    sm = cmeans()
    wd.cell(this, 2, sp[0])
    mst = 3
    for m in sm[0]:
        wd.cell(this, mst, m)
        mst += 1
    this += 1
    if len(sp)>1:
        for i in range(1, len(sp)):
            mst = 3
            wd.insert_rows(this)
            wd.cell(this, 2, sp[i])
            for m in sm[i]:
                wd.cell(this, mst, m)
                mst += 1
            this += 1
    print(time, end = " ")
wdb.save("wb.xlsx")

