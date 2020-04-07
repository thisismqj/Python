import tkinter as tk
from openpyxl import load_workbook
import random

global column
column = 0
wd = load_workbook("wb.xlsx")["Base"]
global word
word = ""
global speech
speech = ""
global mean
mean = ""
global answer
answer = ""
global result
result = False
START = 2297
END = 2374

def flush():
    global word
    global speech
    global mean
    global column
    global answer
    column = random.randint(START, END)
    speech = wd.cell(column, 2).value
    mean = wd.cell(column, 3).value
    if wd.cell(column, 1).value:
        word = wd.cell(column, 1).value
    else:
        while not wd.cell(column, 1).value:
            column -= 1
        word = wd.cell(column, 1).value
    if mode.get() == 1:
        t.set(speech+" "+mean)
        answer = word
    if mode.get() == 2:
        t.set(word+" "+mean)
        answer = speech
    if mode.get() == 3:
        t.set(word+" "+speech)
        answer = mean
    e1.delete(0, END)
def mode_select():
    flush()
def next_(event=0):
    check()
    flush()
def check():
    if e1.get() == answer:
        t1.set("正确")
    else:
        t1.set("错误 正确答案:"+answer)
winmain = tk.Tk()
winmain.title("tk Demo")
fl = tk.Frame(winmain)
fr = tk.Frame(winmain)
mode = tk.IntVar(master=winmain)
tk.Radiobutton(fl, text="单词", variable=mode, value=1, command=mode_select).pack(side=tk.TOP, fill="x")
tk.Radiobutton(fl, text="词性", variable=mode, value=2, command=mode_select).pack(side=tk.TOP, fill="x")
tk.Radiobutton(fl, text="释义", variable=mode, value=3, command=mode_select).pack(side=tk.TOP, fill="x")
fl.pack(side=tk.LEFT)
t = tk.StringVar(winmain, value="左侧选模式")
l = tk.Label(fr, textvariable=t).grid(row=0)
e1 = tk.Entry(fr)
e1.grid(row=1)
e1.bind("<KeyPress-Return>", next_)
fr.pack(side=tk.RIGHT)
t1 = tk.StringVar(winmain)
l1 = tk.Label(fr, textvariable = t1).grid(row=2)
winmain.mainloop()
